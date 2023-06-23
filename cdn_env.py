# https://towardsdatascience.com/beginners-guide-to-custom-environments-in-openai-s-gym-989371673952
# https://github.com/eriklindernoren/ML-From-Scratch/blob/master/mlfromscratch/reinforcement_learning/deep_q_network.py

# https://github.com/openai/gym/blob/master/gym/envs/atari/atari_env.py
# https://github.com/openai/atari-py/blob/5ca24caf395320ead8c77440fe729cf53804616a/atari_py/ale_python_interface.py

import os
import gym
import time
import copy
from datetime import datetime
from utils import beautify_json
from gym.spaces import Discrete, Dict
from prometheus import PrometheusReader
from perform_action import suggest_change
from openpyxl import Workbook, load_workbook
from nginx_configurator import NginxConfigurator
from constants import TRAINING_EDGE_HOST_NAME, CONFIG_FILE_PATH, CONFIG_KEYS, LOGS_PATH, TIME_INTERVAL_BETWEEN_STEPS
from constants import LOGS_DIR
import sys
import traceback
import shutil

class CdnEnv(gym.Env):
    def __init__(self):
        super(CdnEnv, self).__init__()

        self.prometheus_reader = PrometheusReader(host=TRAINING_EDGE_HOST_NAME)
        self.nginx_configurator = NginxConfigurator(CONFIG_FILE_PATH)
        self._initialize_logging()
        self.server_state_goodness = 0
        
        self.action_set = [
            'noop',
            'worker_processes__increase',
            'worker_processes__decrease',
            'worker_connections__increase',
            'worker_connections__decrease',
            'multi_accept__toggle',
            'proxy_cache_min_uses__increase',
            'proxy_cache_min_uses__decrease',
            'output_buffers/1__increase',
            'output_buffers/1__decrease',
            'output_buffers/2__increase',
            'output_buffers/2__decrease',
            'keepalive_timeout__increase',
            'keepalive_timeout__decrease',
            'reset_timedout_connection__toggle',
            'send_timeout__increase',
            'send_timeout__decrease',
            'gzip__toggle',
            'gzip_comp_level__increase',
            'gzip_comp_level__decrease',
            'open_file_cache/max__increase',
            'open_file_cache/max__decrease',
            'open_file_cache/inactive__increase',
            'open_file_cache/inactive__decrease',
            'open_file_cache_valid__increase',
            'open_file_cache_valid__decrease',
            'open_file_cache_min_uses__increase',
            'open_file_cache_min_uses__decrease'
        ]

        self.observation_set = {
            # resources
            #'total_ram_memory': [2, 4, 5, 16, 32, 63],
            #'total_disk': [4, 8, 16, 32, 36],
            #'cpu_cores': [2, 4, 8, 16, 24, 32],
            
            # configuration
            'worker_processes': list(range(1, 16+1)) + ['auto'],
            'worker_connections': [512, 1024, 2048, 4096, 8192],
            'multi_accept': ['off', 'on'],
            'proxy_cache_min_uses': list(range(1, 8+1)),
            'output_buffers/1': list(range(1, 4+1)),
            'output_buffers/2': ['4k', '8k', '16k', '32k', '64k', '128k'],
            'keepalive_timeout': [15, 30, 60, 75, 90],
            'reset_timedout_connection': ['off', 'on'],
            'send_timeout': ['15s', '20s', '30s', '45s', '60s', '75s'],
            'gzip': ['off', 'on'],
            'gzip_comp_level': list(range(1, 9+1)),
            'open_file_cache/max': list(range(2000, 20000, 2000)),
            'open_file_cache/inactive': ['30s', '1m', '90s', '2m', '210s', '5m', '450s', '10m'],
            'open_file_cache_valid': ['30s', '1m', '90s', '2m', '210s', '5m', '10m'],
            'open_file_cache_min_uses': list(range(1, 4+1)),
        }

        self.action_space = Discrete(len(self.action_set))

        # TODO: shall we remember previous action as well? so we can undo some bad config change
        # TODO: optimize observation space
        self.observation_space = Dict({
            # resources
            #'total_ram_memory': Discrete(len(self.observation_set['total_ram_memory'])),
            #'total_disk': Discrete(len(self.observation_set['total_disk'])),
            #'cpu_cores': Discrete(len(self.observation_set['cpu_cores'])),

            # status
            'system_load': Discrete(100),
            'cpu_usage': Discrete(100),
            'cpu_iowait': Discrete(10),
            'ram_memory_usage': Discrete(100),
            #'disk_usage': Discrete(100),
            'disk_io_usage': Discrete(100),
            #'disk_iops_reads': Discrete(100),
            'disk_read_usage': Discrete(100),
            #'disk_iops_writes': Discrete(100),
            'disk_write_usage': Discrete(100),
            'disk_reads_mbps': Discrete(100),
            'disk_writes_mbps': Discrete(100),

            # configuration
            'worker_processes': Discrete(len(self.observation_set['worker_processes'])),
            'worker_connections': Discrete(len(self.observation_set['worker_connections'])),
            'multi_accept': Discrete(len(self.observation_set['multi_accept'])),
            'proxy_cache_min_uses': Discrete(len(self.observation_set['proxy_cache_min_uses'])),
            'output_buffers/1': Discrete(len(self.observation_set['output_buffers/1'])),
            'output_buffers/2': Discrete(len(self.observation_set['output_buffers/2'])),
            'keepalive_timeout': Discrete(len(self.observation_set['keepalive_timeout'])),
            'reset_timedout_connection': Discrete(len(self.observation_set['reset_timedout_connection'])),
            'send_timeout': Discrete(len(self.observation_set['send_timeout'])),
            'gzip': Discrete(len(self.observation_set['gzip'])),
            'gzip_comp_level': Discrete(len(self.observation_set['gzip_comp_level'])),
            'open_file_cache/max': Discrete(len(self.observation_set['open_file_cache/max'])),
            'open_file_cache/inactive': Discrete(len(self.observation_set['open_file_cache/inactive'])),
            'open_file_cache_valid': Discrete(len(self.observation_set['open_file_cache_valid'])),
            'open_file_cache_min_uses': Discrete(len(self.observation_set['open_file_cache_min_uses'])),

            # workflow
            'number_of_requests_per_second': Discrete(100),
            'average_response_size_mb': Discrete(100),

            # performance
            'average_request_processing_time_in_seconds': Discrete(100),

            # 'p95_request_processing_time_in_seconds': Discrete(1000),
            'nginx_cache_hit_ratio': Discrete(100),
            'nginx_writing_connections': Discrete(100),
            'average_network_transmit_mbps': Discrete(100),
            'number_of_5xx_errors_per_second': Discrete(10),
        })

    def step(self, act):
        print(f"step function calls: {act} at {datetime.now().replace(microsecond=0)}")
        self.action = self.action_set[act]
        self.action_time = ''
        self.reward = 0
        self.step_counter += 1
        info = {}

        # TODO: a huge change in input -> another 1-5 minutes delay?
        # because user peaks should not make us think configuration change has been cause of changes
        # change in input has caused the change in observation

        # TODO: should we set a negative reward for 'noop' action? seems NO
        if self.action != 'noop':
            config_name, config_value, valid_action = suggest_change(self.action, self.observation_space, self.observation)
            if not valid_action:
                self.reward = -101
                self.action_time = 'not valid action'
                self._log()
                return self.observation, self.reward, self._check_done(), info

            # map index to value
            if not config_name in self.observation_set.keys():
                print(f"config {config_name} not in self.observation_set.keys()")
                self.reward = -102
                self.action_time = f'config {config_name} not in observation_set'
                self._log()
                return self.observation, self.reward, self._check_done(), info

            config_value = self.observation_set[config_name][config_value]
            self.observation[config_name] = config_value
            self.nginx_configurator.set_config_value(config_name, config_value)
            self.send_configs_to_remote_server()
            self.action_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # TODO how much shall we wait here?
        # wait for a few minutes
        time.sleep(TIME_INTERVAL_BETWEEN_STEPS)

        try:
            self.observation = self._get_observation()
        except Exception:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
            exception_message = ''.join('!! ' + line for line in lines)  # Log it or whatever here
            self._log_to_file(exception_message)

        new_goodness = self._calculate_server_state_goodness()
        if self.step_counter == 1:     # in first iteration, self.server_state_goodness == 0
            self.reward = 0
        else:
            self.reward = new_goodness - self.server_state_goodness
        self.server_state_goodness = new_goodness

        self.render()
        self._log()

        # TODO define termination rule, if any
        # if there has been noop action for some consecutive iterations, done = True

        return self.observation, self.reward, self._check_done(), info

    def reset(self):
        self.observation = self._get_observation()
        self.step_counter = 0
        return self.observation

    def render(self, mode="human"):
        print('*' * 50)
        print('step#:', self.step_counter)
        print('action: ', self.action)
        print('action_time: ', self.action_time)
        print('reward: ', self.reward)
        print('server state goodness:', self.server_state_goodness)

        #mapped_observation = self._map_observation(self.observation)
        #print('mapped observation: ', beautify_json(mapped_observation))

    def _log(self):
        try:
            wb = load_workbook(LOGS_PATH)
            ws1 = wb['Transformed Data']
            ws2 = wb['Original Data']
            self._write_to_worksheet(wb, ws1, self._map_observation(self.observation))
            self._write_to_worksheet(wb, ws2, self._map_observation(self.original_observation))
        except Exception:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
            exception_message = ''.join('!! ' + line for line in lines)  # Log it or whatever here
            print(f"exception in _log: {exception_message}")
            
    def _log_to_file(self, message):
        with open(f"./{LOGS_DIR}/logs.txt", 'a', encoding='utf-8') as file:
            file.write(f"{self.step_counter}\n")
            file.write(f"time: {datetime.now().replace(microsecond=0)}\n")
            file.write(f"action: {self.action}\n")
            file.write(f"action_time: {self.action_time}\n")
            file.write(f"reward: {self.reward}\n")
            file.write(f"server state goodness: {self.server_state_goodness}\n")
            file.write(f"{message}\n")
            file.write('*' * 100 + "\n")

    def _write_to_worksheet(self, workbook, worksheet, mapped_observation):
        worksheet.cell(row=self.step_counter + 1, column=1, value=self.step_counter)
        worksheet.cell(row=self.step_counter + 1, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        worksheet.cell(row=self.step_counter + 1, column=3, value=self.action)
        worksheet.cell(row=self.step_counter + 1, column=4, value=self.action_time)
        worksheet.cell(row=self.step_counter + 1, column=5, value=self.reward)
        worksheet.cell(row=self.step_counter + 1, column=6, value=self.server_state_goodness)

        col = 7
        for key, value in mapped_observation.items():
            worksheet.cell(row=self.step_counter + 1, column=col, value=value)
            col += 1

        workbook.save(LOGS_PATH)

    @staticmethod
    def _initialize_logging():
        if os.path.isfile(LOGS_PATH):
            unique_id = datetime.now().replace(microsecond=0).isoformat(sep='_').replace(':','-')
            shutil.move(LOGS_PATH, os.path.dirname(LOGS_PATH) + '/' + unique_id + '.xlsx')

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Transformed Data"
        ws2 = wb.create_sheet(title="Original Data")
        columns = ["step#", "time", "action", "action_time", "reward", "server state goodness", 
                   #"total_ram_memory", "total_disk", "cpu_cores",
                   "system_load", "cpu_usage", "cpu_iowait", "ram_memory_usage",
                   #"disk_usage",
                   "disc_io_usage", "disc_read_usage", "disc_write_usage", "disc_reads_mbps", "disc_writes_mbps",
                   "number_of_requests_per_second", "average_response_size_mb",
                   "average_request_processing_time_in_seconds",
                   "nginx_cache_hit_ratio", "nginx_writing_connections", "average_network_transmit_mbps",
                   "number_of_5xx_errors_per_second", "worker_processes", "worker_connections", "multi_accept",
                   "proxy_cache_min_uses", "output_buffers/1", "output_buffers/2",
                   "keepalive_timeout",  "reset_timedout_connection", "send_timeout", "gzip", "gzip_comp_level",
                   "open_file_cache/max", "open_file_cache/inactive", "open_file_cache_valid",
                   "open_file_cache_min_uses"]

        for index, col in enumerate(columns):
            ws1.cell(row=1, column=index + 1, value=col)
            ws2.cell(row=1, column=index + 1, value=col)

        wb.save(LOGS_PATH)

    def _map_observation(self, observation):
        try:
            mapped_observation = copy.deepcopy(observation)
            for key, value in observation.items():
                if key in self.observation_set.keys():
                    mapped_observation[key] = self.observation_set[key][value]
            return mapped_observation
        except Exception:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
            exception_message = ''.join('!! ' + line for line in lines)  # Log it or whatever here
            print(f"exception in _map_observation: {exception_message}")
            print(f"_map_observation: key:{key}, value:{value}")
            

    def _get_observation(self):
        print(f'_get_observation at {datetime.now().replace(microsecond=0)}')
        observation = self.prometheus_reader.read_environment()
        observation.update(self.nginx_configurator.get_config_values(CONFIG_KEYS))
        print(beautify_json(observation))
        self.original_observation = copy.deepcopy(observation)

        for key, value in observation.items():
            #print(f"observation.item {key}: {value}")
            # map value to index
            if key in self.observation_set.keys():
                possible_values = self.observation_set[key]
                index = possible_values.index(value) if value in possible_values else possible_values[-1]
                observation[key] = index
                self.original_observation[key] = index
            else:
                # the idea behind this action is that there is no big difference between close numbers
                # in our observation space
                new_value = value
                if new_value >= self.observation_space[key].n:
                    # if value is not valid change it to the last eligible value
                    new_value = self.observation_space[key].n - 1
                observation[key] = new_value

        return observation

    # Estimate state 'goodness' based on system_performance
    # We define goodness based on better system status and better system performance
    def _calculate_server_state_goodness(self):
        #return previous server state goodness in case of no requests
        if self.observation['number_of_requests_per_second'] == 0:
            return self.server_state_goodness   
        goodness = 0

        # TODO: check weights
        weights = {
            'system_load': 10,
            'cpu_usage': 10,
            'cpu_iowait': 100,
            'ram_memory_usage': 100,
            'disk_io_usage': 10,
            'disk_read_usage': 10,
            'disk_write_usage': 10,
            'disk_reads_mbps': 10,
            'disk_writes_mbps': 10,
            'average_request_processing_time_in_seconds': 100,
            'nginx_cache_hit_ratio': 10,
            'nginx_writing_connections': 1,
            'average_network_transmit_mbps': 1,
            'number_of_5xx_errors_per_second': 100,
        }

        goodness += weights['system_load'] * (
                100 - self.observation['system_load'])
        goodness += weights['cpu_usage'] * (
                100 - self.observation['cpu_usage'])
        goodness += weights['cpu_iowait'] * (
                100 - self.observation['cpu_iowait'])
        goodness += weights['ram_memory_usage'] * (100 - self.observation['ram_memory_usage'])
        goodness += weights['disk_io_usage'] * (100 - self.observation['disk_io_usage'])
        goodness += weights['disk_read_usage'] * (100 - self.observation['disk_read_usage'])
        goodness += weights['disk_write_usage'] * (100 - self.observation['disk_write_usage'])
        goodness += weights['disk_reads_mbps'] * (100 - self.observation['disk_reads_mbps'])
        goodness += weights['disk_writes_mbps'] * (100 - self.observation['disk_writes_mbps'])

        goodness += weights['average_request_processing_time_in_seconds'] * (
                    100 - self.observation['average_request_processing_time_in_seconds'])
        goodness += weights['nginx_cache_hit_ratio'] * (self.observation['nginx_cache_hit_ratio'])
        goodness += weights['nginx_writing_connections'] * (self.observation['nginx_writing_connections'])
        goodness += weights['average_network_transmit_mbps'] * (self.observation['average_network_transmit_mbps'])
        goodness += weights['number_of_5xx_errors_per_second'] * (100 - self.observation['number_of_5xx_errors_per_second'])

        return goodness

    def _check_done(self):
        if self.step_counter >= 20:
            return True
        else:
            return False

    @staticmethod
    def send_configs_to_remote_server():
        os.system(f'scp -P 24900 -r ./nginx-configs/* vahid@{TRAINING_EDGE_HOST_NAME}.myket.ir:/home/vahid/nginx-configs && ssh -p 24900 '
                  f'vahid@{TRAINING_EDGE_HOST_NAME}.myket.ir "sudo nginx -c /home/vahid/nginx-configs/nginx.conf -t && sudo systemctl reload nginx" &>> '
                  'config_sender_logs.txt')